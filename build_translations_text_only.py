"""
Generate translations_text_only.xlsx — one sheet per HTML page,
columns: EN | DE  (no keys)
"""
import re
import json
import glob
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE = Path(__file__).parent

# ── 1. Parse i18n.js ────────────────────────────────────────────────────────
i18n_text = (BASE / "assets" / "i18n.js").read_text(encoding="utf-8")

def extract_lang_block(text, lang):
    pattern = rf'{lang}:\s*\{{(.*?)\n  \}}'
    m = re.search(pattern, text, re.DOTALL)
    if not m:
        return {}
    block = "{" + m.group(1) + "}"
    block = re.sub(r',\s*\}', '}', block)
    try:
        return json.loads(block)
    except json.JSONDecodeError:
        result = {}
        for line in m.group(1).splitlines():
            lm = re.match(r'\s*"([^"]+)":\s*"(.*)",?\s*$', line)
            if lm:
                result[lm.group(1)] = lm.group(2)
        return result

en = extract_lang_block(i18n_text, "en")
de = extract_lang_block(i18n_text, "de")

# ── 2. Scan each HTML for data-i18n* keys ───────────────────────────────────
KEY_RE = re.compile(r'data-i18n(?:-html|-placeholder)?\s*=\s*"([^"]+)"')

def page_keys(html_path):
    text = Path(html_path).read_text(encoding="utf-8")
    seen = set()
    keys = []
    for m in KEY_RE.finditer(text):
        k = m.group(1)
        if k not in seen:
            seen.add(k)
            keys.append(k)
    return keys

html_files = sorted(glob.glob(str(BASE / "*.html")))

# ── 3. Build Excel ───────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
wb.remove(wb.active)

HEADER_FILL = PatternFill("solid", fgColor="1A3557")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
CELL_FONT   = Font(name="Calibri", size=10)
ALIGN_WRAP  = Alignment(wrap_text=True, vertical="top")
THIN        = Side(border_style="thin", color="CCCCCC")
BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def safe_sheet_name(filename):
    name = Path(filename).stem[:31]
    for c in r'\/?*[]':
        name = name.replace(c, "-")
    return name

for html_path in html_files:
    sheet_name = safe_sheet_name(html_path)
    ws = wb.create_sheet(title=sheet_name)

    ws.column_dimensions["A"].width = 70
    ws.column_dimensions["B"].width = 70

    for col, label in enumerate(["EN", "DE"], start=1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
    ws.row_dimensions[1].height = 20

    keys = page_keys(html_path)
    if not keys:
        ws.cell(row=2, column=1, value="(no translatable text found)")
        continue

    for row_idx, key in enumerate(keys, start=2):
        en_cell = ws.cell(row=row_idx, column=1, value=en.get(key, ""))
        en_cell.font = CELL_FONT
        en_cell.border = BORDER
        en_cell.alignment = ALIGN_WRAP

        de_cell = ws.cell(row=row_idx, column=2, value=de.get(key, ""))
        de_cell.font = CELL_FONT
        de_cell.border = BORDER
        de_cell.alignment = ALIGN_WRAP

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:B{len(keys)+1}"

out_path = BASE / "translations_text_only.xlsx"
wb.save(out_path)
print(f"Saved: {out_path}  ({len(html_files)} sheets)")
