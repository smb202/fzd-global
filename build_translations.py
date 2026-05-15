"""
Generate translations.xlsx — one sheet per HTML page,
columns: Key | EN | DE
"""
import re
import json
import glob
import os
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = Path(__file__).parent

# ── 1. Parse i18n.js → {lang: {key: value}} ────────────────────────────────
i18n_text = (BASE / "assets" / "i18n.js").read_text(encoding="utf-8")

def extract_lang_block(text, lang):
    """Pull out the en:{...} or de:{...} object as a Python dict."""
    pattern = rf'{lang}:\s*\{{(.*?)\n  \}}'
    m = re.search(pattern, text, re.DOTALL)
    if not m:
        return {}
    block = "{" + m.group(1) + "}"
    # convert JS object to JSON: keys need double-quotes (they already have them)
    # handle trailing commas
    block = re.sub(r',\s*\}', '}', block)
    try:
        return json.loads(block)
    except json.JSONDecodeError:
        # fallback: parse line by line
        result = {}
        for line in m.group(1).splitlines():
            lm = re.match(r'\s*"([^"]+)":\s*"(.*)",?\s*$', line)
            if lm:
                result[lm.group(1)] = lm.group(2)
        return result

en = extract_lang_block(i18n_text, "en")
de = extract_lang_block(i18n_text, "de")

# ── 2. Scan each HTML for data-i18n* keys (preserving order) ────────────────
KEY_ATTRS = ["data-i18n", "data-i18n-html", "data-i18n-placeholder"]
KEY_RE = re.compile(r'data-i18n(?:-html|-placeholder)?\s*=\s*"([^"]+)"')

def page_keys(html_path):
    """Return ordered list of unique keys used on this page."""
    text = Path(html_path).read_text(encoding="utf-8")
    seen = set()
    keys = []
    for m in KEY_RE.finditer(text):
        k = m.group(1)
        if k not in seen:
            seen.add(k)
            keys.append(k)
    return keys

html_files = sorted(glob.glob(str(BASE / "*.html")))

# ── 3. Build Excel ───────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
wb.remove(wb.active)  # remove default sheet

# Styles
HEADER_FILL = PatternFill("solid", fgColor="1A3557")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
KEY_FILL    = PatternFill("solid", fgColor="EEF2F7")
KEY_FONT    = Font(color="555555", name="Calibri", size=10, italic=True)
CELL_FONT   = Font(name="Calibri", size=10)
CELL_FONT_DE = Font(name="Calibri", size=10)
ALIGN_WRAP  = Alignment(wrap_text=True, vertical="top")
THIN        = Side(border_style="thin", color="CCCCCC")
BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def safe_sheet_name(filename):
    name = Path(filename).stem          # e.g. "industry-aerospace"
    name = name[:31]                    # Excel 31-char limit
    for c in r'\/?*[]':
        name = name.replace(c, "-")
    return name

for html_path in html_files:
    sheet_name = safe_sheet_name(html_path)
    ws = wb.create_sheet(title=sheet_name)

    # Header row
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 60

    for col, label in enumerate(["Key", "EN", "DE"], start=1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
    ws.row_dimensions[1].height = 20

    keys = page_keys(html_path)
    if not keys:
        ws.cell(row=2, column=1, value="(no translatable text found)")
        continue

    for row_idx, key in enumerate(keys, start=2):
        en_val = en.get(key, "")
        de_val = de.get(key, "")

        key_cell = ws.cell(row=row_idx, column=1, value=key)
        key_cell.font = KEY_FONT
        key_cell.fill = KEY_FILL
        key_cell.border = BORDER
        key_cell.alignment = ALIGN_WRAP

        en_cell = ws.cell(row=row_idx, column=2, value=en_val)
        en_cell.font = CELL_FONT
        en_cell.border = BORDER
        en_cell.alignment = ALIGN_WRAP

        de_cell = ws.cell(row=row_idx, column=3, value=de_val)
        de_cell.font = CELL_FONT_DE
        de_cell.border = BORDER
        de_cell.alignment = ALIGN_WRAP

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:C{len(keys)+1}"

out_path = BASE / "translations.xlsx"
wb.save(out_path)
print(f"Saved: {out_path}  ({len(html_files)} sheets)")
